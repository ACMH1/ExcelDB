# -*- coding: utf-8 -*-
"""
Created on Sat Mar 31 00:45:43 2018

@author: User
"""

import unittest
import openpyxl
import ExcelFactory

class TestSheetDAOImpl(unittest.TestCase):
    
    @classmethod
    def setUpClass(cls):
        cls.file_name = "TestSheetDAOImpl.xlsx"
        cls.workbook = openpyxl.Workbook()
        title_array = ["alpha", "beta", "charlie"]
        for title in title_array:
            cls.workbook.create_sheet(title)
        cls.beta_headers_array = ["header1", "header2", "header3"]
        for x, header in enumerate(cls.beta_headers_array):
            cls.workbook["beta"].cell(column = x+1, row = 1).value = header
        cls.beta_workbook = cls.workbook["beta"]
        cls.workbook.save(cls.file_name)
    
    @classmethod
    def tearDownClass(cls):
        cls.workbook.save(cls.file_name)
    
    def setUp(self):
        pass
    
    def tearDown(self):
        pass
    
    def test_SheetDAOImpl_init(self):
        self.assertIsInstance(ExcelFactory.SheetDAOImpl(self.beta_workbook), ExcelFactory.SheetDAOImpl)
        self.assertEqual(ExcelFactory.SheetDAOImpl(self.beta_workbook).column_headers, self.beta_headers_array)
    
    def test_create_row(self):
        pass
    
    def test_find_row(self):
        pass
    
    def test_update_row(self):
        pass
    
    def test_delete_row(self):
        pass
    
if __name__ == "__main__":
    unittest.main()