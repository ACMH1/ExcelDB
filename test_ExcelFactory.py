# -*- coding: utf-8 -*-
"""
Created on Sun Mar 25 03:33:12 2018

@author: User
"""

import unittest
import openpyxl
import ExcelFactory
#import os

class TestExcelFactory(unittest.TestCase):
    
    @classmethod
    def setUpClass(cls):
        cls.file_name = "TestExcelFactory.xlsx"
        cls.workbook = openpyxl.Workbook()
        title_array = ["alpha", "beta", "charlie"]
        for title in title_array:
            cls.workbook.create_sheet(title)
        cls.beta_headers_array = ["header1", "header2", "header3"]
        for x, header in enumerate(cls.beta_headers_array):
            cls.workbook["beta"].cell(column = x+1, row = 1).value = header
        cls.workbook.save(cls.file_name)
        cls.ExcelDAO = ExcelFactory.SheetDAOFactory(cls.file_name)
#        print("setup test case")
    
    @classmethod
    def tearDownClass(cls):
        cls.workbook.save(cls.file_name)
#        os.remove(cls.file_name)
#        print("teardown test case")
    
#    def setUp(self):
#        print("set up test")
    
#    def tearDown(self):
#        self.workbook.save(self.file_name)
#        print("teardown test")
    
    def test_ExcelFactory_init(self):
        
        self.assertIsInstance(ExcelFactory.SheetDAOFactory(self.file_name), ExcelFactory.SheetDAOFactory)
        self.assertIsInstance(ExcelFactory.SheetDAOFactory("xyz.xlsx"), ExcelFactory.SheetDAOFactory)
        
#        with self.assertRaises(FileNotFoundError):
#            ExcelFactory.SheetDAOFactory("xyz.xlsx")

    def test_get_sheet_DAO(self):
        
         self.assertIsInstance(self.ExcelDAO.get_sheet_DAO("beta"), ExcelFactory.SheetDAOImpl)
         self.assertEqual(self.ExcelDAO.get_sheet_DAO("beta").column_headers, self.beta_headers_array)
    
    def test_create_sheet_DAO(self):
        
        delta_headers_array = ["header4", "header5", "header6"]
        self.assertIsInstance(self.ExcelDAO.create_sheet_DAO("delta", delta_headers_array), ExcelFactory.SheetDAOImpl)
        self.assertEqual(self.ExcelDAO.create_sheet_DAO("delta", delta_headers_array).column_headers, delta_headers_array)
        
    def test_save(self):
        
        self.ExcelDAO.workbook["alpha"]["A1"] = 10
        self.ExcelDAO.save()
        testFactory = ExcelFactory.SheetDAOFactory(self.file_name)
        self.assertEqual(testFactory.workbook["alpha"]["A1"], 10)
    
if __name__ == "__main__":
    unittest.main()