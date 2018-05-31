# -*- coding: utf-8 -*-
"""
Created on Sat Mar 10 17:45:32 2018

@author: User
"""

import openpyxl

class SheetDAOFactory:

    def __init__(self, file_name):
        self.file_name = file_name
        try:
            self.workbook = openpyxl.load_workbook(file_name)
            self.sheet_list = self.workbook.get_sheet_names()
            self.schema = self.workbook["Schema"]
            self.schema_headers = [column[0].value for column in self.schema.iter_cols()]
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.schema = self.workbook.active
            self.schema.title = "Schema"
            self.schema_headers = ["Sheet", "Order", "Name", "Key", "Unique", "Foreign Sheet", "Foreign Key"]
            for n, header in enumerate(self.schema_headers):
                self.schema.cell(row=1, column = n + 1).value = header

    def save(self):
        self.workbook.save(self.file_name)

    def create_sheet_DAO(self, sheet_name, headers):
        worksheet = self.workbook.create_sheet(sheet_name)
        for n, header_name in enumerate(headers):
            worksheet.cell(row=1, column=n + 1).value = header_name
            
            last_row_num = self.schema.max_row + 1
            self.schema.cell(row=last_row_num, column = self.schema_headers.index("Sheet") + 1).value = sheet_name
            self.schema.cell(row=last_row_num, column = self.schema_headers.index("Order") + 1).value = n
            self.schema.cell(row=last_row_num, column = self.schema_headers.index("Name") + 1).value = header_name
            
        return SheetDAOImpl(worksheet, headers=headers)

    def get_sheet_DAO(self, sheet_name):
        headers = [row[self.schema_headers.index("Name")].value for row in self.schema.iter_rows() if row[self.schema_headers.index("Sheet")].value == sheet_name]
        return SheetDAOImpl(self.workbook[sheet_name], headers=headers)

class SheetDAOImpl:

    def __init__(self, worksheet, headers=None):
        self.worksheet = worksheet
        if headers is None:
            self.column_headers = []
            for column in self.worksheet.iter_cols():
                self.column_headers.append(column[0].value)
        else:
            self.column_headers = headers

    def create_row(self, row):
        last_row_num = self.worksheet.max_row + 1
        for prop in row.keys():
            self.worksheet.cell(row=last_row_num, column=self.column_headers.index(prop) + 1).value = row[prop]

    def find_row(self, row):
        pass

    def update_row(self, row):
        pass

    def delete_row(self, row):
        pass
